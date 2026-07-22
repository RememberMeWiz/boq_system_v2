"""
Plan2Takeoff V2 — DUPA (Detailed Unit Price Analysis) Rate Extractor & QA Engine
Parses official DPWH Detailed Unit Price Analysis Excel workbooks stored in
backend/reference_data/dupa_files/ for QA rate verification & fallback unit costs.
"""

import os
import openpyxl
from typing import Dict, Any, List, Optional

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF_DATA_DIR = os.path.join(BASE_DIR, "reference_data")
DUPA_DIR = os.path.join(REF_DATA_DIR, "dupa_files")


class DUPARateLoader:
    def __init__(self):
        self.residential_file = os.path.join(DUPA_DIR, "6. Detailed Unit Price Analysis (DUPA) -Residential Projects.xlsx")
        self.roads_file = os.path.join(DUPA_DIR, "7. Detailed Unit Price Analysis (DUPA) - Roads Projects.xlsx")
        self._cache = None

    def load_rates(self) -> Dict[str, Any]:
        """Parses DUPA Excel files and returns a dictionary of Pay Item unit costs."""
        if self._cache:
            return self._cache

        rates = {}

        if os.path.exists(self.residential_file):
            try:
                wb = openpyxl.load_workbook(self.residential_file, data_only=True)
                for sheetname in wb.sheetnames:
                    if sheetname in ('INPUT DATA', 'BOE', 'ABC'):
                        continue

                    ws = wb[sheetname]
                    # Extract pay item title, unit cost, labor, material, equipment
                    item_code = sheetname.split()[0]
                    description = sheetname

                    mat_cost = 0.0
                    lab_cost = 0.0
                    eqp_cost = 0.0
                    total_unit_cost = 0.0

                    # Scan sheet for totals
                    for row in ws.iter_rows(values_only=True):
                        if not row or len(row) < 3:
                            continue
                        r_str = " ".join([str(c) for c in row if c is not None]).upper()
                        
                        if "DIRECT UNIT COST" in r_str or "TOTAL DIRECT COST" in r_str or "UNIT COST" in r_str:
                            for val in reversed(row):
                                if isinstance(val, (int, float)) and val > 0:
                                    total_unit_cost = float(val)
                                    break
                        elif "TOTAL MATERIAL" in r_str:
                            for val in reversed(row):
                                if isinstance(val, (int, float)) and val > 0:
                                    mat_cost = float(val)
                                    break
                        elif "TOTAL LABOR" in r_str:
                            for val in reversed(row):
                                if isinstance(val, (int, float)) and val > 0:
                                    lab_cost = float(val)
                                    break
                        elif "TOTAL EQUIPMENT" in r_str:
                            for val in reversed(row):
                                if isinstance(val, (int, float)) and val > 0:
                                    eqp_cost = float(val)
                                    break

                    if total_unit_cost > 0 or (mat_cost + lab_cost + eqp_cost) > 0:
                        total = total_unit_cost or (mat_cost + lab_cost + eqp_cost)
                        rates[sheetname] = {
                            "item_code": item_code,
                            "description": description,
                            "material_unit_cost": round(mat_cost, 2),
                            "labor_unit_cost": round(lab_cost, 2),
                            "equipment_unit_cost": round(eqp_cost, 2),
                            "total_unit_cost": round(total, 2),
                            "source": "DPWH DUPA Residential"
                        }
                wb.close()
            except Exception as e:
                print(f"[DUPA Loader] Error loading residential rates: {e}")

        self._cache = rates
        return rates


def get_dupa_qa_summary() -> Dict[str, Any]:
    """Returns a QA summary of available DUPA unit rates for verification."""
    loader = DUPARateLoader()
    rates = loader.load_rates()
    return {
        "status": "active",
        "total_dupa_items": len(rates),
        "items": list(rates.keys())
    }
